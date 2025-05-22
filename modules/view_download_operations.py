# modules/view_download_operations.py

import os
import logging
from io import BytesIO
from datetime import datetime
import requests
from openpyxl import Workbook, load_workbook
from modules.tableau_operations import get_tableau_connection, sign_out_tableau

# Updated list of consolidated views
# You can group them under a single "workbook" key since they're now all standalone views.
_WORKBOOK_VIEWS = {
    'combined_views': [
        ('Main_Combo',          'c9cb3cd4-c259-4d49-a23c-439e1f0bd1c5'),
        ('Stragglers',          '7e0839f6-373d-46ca-977e-b75d06e78b6b'),
        ('Basin_Crew_Count',    '61e30382-41e6-4a3e-a465-c402450d0057'),
        ('PnL_CAN_GL',          '647bd58f-3898-4479-b2e8-499bdb9ac248'),
        ('GL_Basin_Pivot',      '50b6f916-3629-4e25-b400-4222a972bd39'),
        ('Project_VM',          '18667dc0-a4d0-48d4-8fb6-be5a8ef2601c'),
        ('Transload',           '77ffa2c2-b950-49ba-a93e-d4353da2d7d7'),
        ('Unassigned_Rev_CA',   '2881be67-415a-4250-aa3b-e0340741ad9a'),
        ('Unalloc_Costs',       '0e1536da-2a9f-4a75-aa68-2493f3694f2a'),
        ('FCAST',               '7a2e640c-5147-4ecf-89b4-2795ad7ab000'),
        ('Project_List',        '64e986a7-b612-495d-9c8e-77a3be722ba8'),
    ]
}

def download_all_views(config_parser, save_dir):
    """
    Signs in once, fetches all configured views as crosstab/excel,
    appends each to its own sheet, then signs out and saves one workbook.
    Returns the full path of the saved file.
    """
    conn    = get_tableau_connection(config_parser)
    site_id = conn.site_id
    logging.info(f"Downloading views for site_id: {site_id}")

    wb = Workbook()
    wb.remove(wb.active)

    tab_cfg = config_parser["tableau_online"]
    for wb_id, views in _WORKBOOK_VIEWS.items():
        logging.info(f"Workbook '{wb_id}' contains {len(views)} views")
        for view_name, view_id in views:
            sheet_name = view_name[:31]  # Excel sheet names limited to 31 chars
            orig       = sheet_name
            count      = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{orig}_{count}"
                count += 1

            try:
                url = (
                    f"{tab_cfg['server']}/api/{conn.api_version}"
                    f"/sites/{site_id}/views/{view_id}/crosstab/excel"
                )
                headers = {"X-Tableau-Auth": conn.auth_token}
                resp    = requests.get(url, headers=headers)
                resp.raise_for_status()

                tmp = load_workbook(filename=BytesIO(resp.content), data_only=True).active
                tgt = wb.create_sheet(sheet_name)
                for row in tmp.iter_rows(values_only=True):
                    tgt.append(row)

                logging.info(f"Fetched view '{view_name}' -> sheet '{sheet_name}'")
            except Exception as e:
                logging.error(f"Failed to fetch '{view_name}': {e}")
                continue

    sign_out_tableau(conn)

    ts    = datetime.now().strftime("%Y%m%d%H%M%S")
    fname = f"Tableau_Exports_{ts}.xlsx"
    out   = os.path.join(save_dir, fname)
    wb.save(out)
    logging.info(f"All views saved to {out}")
    return out

"""
modules/comparison_operations.py

Handles merging PnL_CAN_GL & Data Sort CAD into the final "Comparison" sheet.
"""

import sys
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import messagebox

def create_comparison_sheet(new_workbook_path, latest_month, ref_table_path):
    """
    Merges 'PnL_CAN_GL' + 'Data Sort CAD', references 'Account Groups' in ref_table_path
    => "Comparison" sheet.
    """
    logging.info("Creating Comparison sheet.")
    try:
        wb = load_workbook(new_workbook_path)
        if "PnL_CAN_GL" not in wb.sheetnames:
            raise ValueError("Sheet 'PnL_CAN_GL' not found in the new workbook.")
        if "Data Sort CAD" not in wb.sheetnames:
            raise ValueError("Sheet 'Data Sort CAD' not found in the new workbook.")

        pnl_sheet = wb["PnL_CAN_GL"]
        data_sort_sheet = wb["Data Sort CAD"]

        # Read them into DataFrames
        pnl_df = pd.DataFrame(pnl_sheet.values)
        pnl_df.columns = pnl_df.iloc[0]
        pnl_df = pnl_df[1:]
        logging.info(f"PnL_CAN_GL columns: {pnl_df.columns.tolist()}")

        # Identify last col as the "pnl_latest_month"
        pnl_latest_month = pnl_df.columns[-1]
        if pnl_latest_month not in pnl_df.columns:
            sys.exit(f"Latest month col '{pnl_latest_month}' not found in PnL_CAN_GL.")

        # Convert to float => int
        pnl_df[pnl_latest_month] = (
            pnl_df[pnl_latest_month]
            .replace({r'\$': '', r'\(': '-', r'\)': ''}, regex=True)
            .astype(float)
            .round(0)
            .astype(int)
        )

        data_sort_df = pd.DataFrame(data_sort_sheet.values)
        data_sort_df.columns = data_sort_df.iloc[0]
        data_sort_df = data_sort_df[1:]
        logging.info(f"Data Sort CAD columns: {data_sort_df.columns.tolist()}")

        if latest_month not in data_sort_df.columns:
            sys.exit(f"Latest month '{latest_month}' not found in Data Sort CAD sheet.")

        data_sort_df[latest_month] = (
            data_sort_df[latest_month]
            .replace({',': '', r'\(': '-', r'\)': ''}, regex=True)
            .astype(float)
        )

        # Ensure 'Account Number' / 'Account2'
        if 'Account Number' not in pnl_df.columns:
            sys.exit("'Account Number' not found in PnL_CAN_GL sheet.")
        pnl_df['Account Number'] = pnl_df['Account Number'].astype(str).str.strip()

        if 'Account2' not in data_sort_df.columns:
            sys.exit("'Account2' not found in Data Sort CAD sheet.")
        data_sort_df['Account2'] = data_sort_df['Account2'].astype(str).str.strip()

        # Read "Account Groups" from ref_table_path
        ref_wb = load_workbook(ref_table_path)
        if "Account Groups" not in ref_wb.sheetnames:
            sys.exit("Worksheet 'Account Groups' not found in ref_table. Please add it.")

        account_groups_df = pd.read_excel(ref_table_path, sheet_name='Account Groups')
        if 'Account Number' not in account_groups_df.columns:
            sys.exit("'Account Number' missing in 'Account Groups' sheet.")
        if 'High CK (group)' not in account_groups_df.columns:
            sys.exit("'High CK (group)' missing in 'Account Groups' sheet.")

        account_groups_df['Account Number'] = account_groups_df['Account Number'].astype(str).str.strip()

        # Merge
        merged_df = pd.merge(
            pnl_df, data_sort_df,
            left_on='Account Number', right_on='Account2',
            how='inner', suffixes=('_PnL', '_CAD')
        )
        logging.info(f"After merge: {merged_df.shape[0]} rows.")

        # Delta
        merged_df['Delta'] = merged_df[f'{latest_month}_CAD'] - merged_df[f'{pnl_latest_month}_PnL']

        # Merge w/ Account Groups
        merged_df = pd.merge(
            merged_df,
            account_groups_df[['Account Number', 'High CK (group)']],
            on='Account Number', how='left'
        )

        missing = merged_df[merged_df['High CK (group)'].isna()]
        if not missing.empty:
            missing_accounts = missing['Account Number'].tolist()
            logging.warning(f"Accounts missing High CK (group): {missing_accounts}")
            messagebox.showwarning(
                "Missing Account Groups",
                f"The following accounts are missing in 'Account Groups':\n{missing_accounts}\nPlease update them."
            )

        # Filter for 'Revenue' / 'Variable Cost' & delta != 0
        final_df = merged_df[
            (merged_df['High CK (group)'].isin(['Revenue', 'Variable Cost'])) &
            (merged_df['Delta'] != 0)
        ]

        req_cols = [
            'Account Number',
            'Account Desc',
            f'{pnl_latest_month}_PnL',
            f'{latest_month}_CAD',
            'Delta'
        ]
        for col in req_cols:
            if col not in final_df.columns:
                sys.exit(f"Missing '{col}' in final data for comparison.")

        final_df = final_df[req_cols].copy()
        final_df.rename(columns={
            'Account Desc': 'Account Desc',
            f'{pnl_latest_month}_PnL': 'PnL_CAN_GL',
            f'{latest_month}_CAD': 'Data_Sort_CAD'
        }, inplace=True)

        # Write to 'Comparison'
        if "Comparison" in wb.sheetnames:
            del wb["Comparison"]
        comp_sheet = wb.create_sheet("Comparison")

        for r in dataframe_to_rows(final_df, index=False, header=True):
            comp_sheet.append(r)

        wb.save(new_workbook_path)
        logging.info("Comparison sheet created & saved.")
        return new_workbook_path
    except Exception as e:
        logging.error(f"Error in create_comparison_sheet: {e}")
        sys.exit("Failed to create Comparison sheet.")

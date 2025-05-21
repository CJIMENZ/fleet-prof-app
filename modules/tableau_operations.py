# modules/tableau_operations.py

import os
import sys
import logging
import requests
from tableau_api_lib import TableauServerConnection
from openpyxl import load_workbook
from urllib import parse

def get_tableau_connection(config_parser):
    """
    Signs in to Tableau and returns a live TableauServerConnection.
    """
    tab_cfg = config_parser["tableau_online"]
    cfg = {
        "tableau_online": {
            "server":                        tab_cfg["server"],
            "api_version":                   tab_cfg["api_version"],
            "personal_access_token_name":    tab_cfg["personal_access_token_name"],
            "personal_access_token_secret":  tab_cfg["personal_access_token_secret"],
            "site_name":                     tab_cfg["site_name"],
            "site_url":                      tab_cfg["site_url"]
        }
    }
    conn = TableauServerConnection(cfg, env="tableau_online")
    conn.sign_in()
    logging.info("Signed in to Tableau Server.")
    return conn

def sign_out_tableau(conn):
    """
    Signs out from an active TableauServerConnection.
    """
    try:
        conn.sign_out()
        logging.info("Signed out from Tableau Server.")
    except Exception as e:
        logging.error(f"Error signing out of Tableau: {e}")

def integrate_tableau(
    config_parser,
    new_workbook_path,
    latest_month,
    view_id="79a33d76-77fd-4c1d-846c-18b02c606d71",
):
    """
    Downloads the PnL_CAN_GL crosstab filtered by ``latest_month`` from Tableau
    and appends it as a sheet named ``PnL_CAN_GL`` in ``new_workbook_path``.
    The ``latest_month`` value is passed via the Tableau parameter
    ``"Year-Month String"`` so the downloaded data matches the FX month.
    """
    logging.info("Integrating Tableau data.")
    try:
        # build the connection
        conn = get_tableau_connection(config_parser)

        # use the already‑scoped site_id
        site_id = conn.site_id
        logging.info(f"Using site_id: {site_id}")

        # fetch the crosstab Excel filtered by the provided month
        field = "Year-Month String"
        field_q = parse.quote_plus(field)
        value_q = parse.quote_plus(str(latest_month))
        endpoint = (
            f"/api/{conn.api_version}/sites/{site_id}/views/{view_id}/crosstab/excel?vf_{field_q}={value_q}"
        )
        full_url = conn.server + endpoint
        headers = {"X-Tableau-Auth": conn.auth_token}

        resp = requests.get(full_url, headers=headers)
        resp.raise_for_status()

        # save to temp file
        excel_file = os.path.join(os.path.dirname(new_workbook_path), "PnL_CAN_GL.xlsx")
        with open(excel_file, "wb") as f:
            f.write(resp.content)

        sign_out_tableau(conn)

        # copy into the target workbook
        pnl_wb    = load_workbook(excel_file, data_only=True)
        pnl_sheet = pnl_wb.active

        wb = load_workbook(new_workbook_path)
        if "PnL_CAN_GL" in wb.sheetnames:
            del wb["PnL_CAN_GL"]
        wb.create_sheet("PnL_CAN_GL")
        target = wb["PnL_CAN_GL"]

        for row in pnl_sheet.iter_rows(values_only=True):
            target.append(row)

        wb.save(new_workbook_path)
        logging.info("PnL_CAN_GL sheet added to the main workbook.")
        return new_workbook_path

    except Exception as e:
        logging.error(f"Error in integrate_tableau: {e}")
        sys.exit("Failed to integrate Tableau data.")

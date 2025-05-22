# modules/cks_pivot_operations.py

import os
import sys
import logging
import warnings
import calendar
from datetime import datetime, date
from openpyxl import load_workbook, Workbook

# suppress openpyxl date‑serial warnings
warnings.filterwarnings(
    "ignore",
    message="Cell .* is marked as a date but the serial value .*",
    category=UserWarning,
    module=r"openpyxl\.worksheet\._reader"
)

def pivot_cks_data_to_ref(finance_file, ref_file, target_month_str):
    """
    Reads 'NIS Details by Basin - US', 'NIS Details by Basin - CA', and now
    'NIS Details by Basin - AU' from finance_file,
    finds the column in row 3 == target_month_str (Mon-YY),
    pivots by scanning column A (basin) & F (data field),
    then appends (with Date/Year/Month/Q/M-Y/Q-Y) to
    ref_file's 'CK data Pivot' sheet.
    """
    logging.info(f"Pivoting CKs data from '{finance_file}' for {target_month_str} -> '{ref_file}'.")

    # 1) Load the finance (CKs) workbook
    try:
        wb_fin = load_workbook(finance_file, data_only=True)
    except FileNotFoundError:
        logging.error(f"File not found: {finance_file}")
        sys.exit("CKs finance file missing.")

    # 2) Verify input sheets for US/CA
    for sheet in ("NIS Details by Basin - US", "NIS Details by Basin - CA"):
        if sheet not in wb_fin.sheetnames:
            logging.error(f"'{sheet}' not in {finance_file}")
            sys.exit(f"Missing sheet {sheet}")

    ws_us = wb_fin["NIS Details by Basin - US"]
    ws_ca = wb_fin["NIS Details by Basin - CA"]

    # 3) Map basins
    dest_us = ["BK","DJ","PR","SJ","OT","AP","PM","EF","HV","MC","Corp","UN"]
    src_us  = ["Williston","DJ","Powder","San Juan","Other",
               "Marcellus Utica","Permian","Eagleford","Haynesville",
               "Midcon","Corporate","Unita"]
    dest_ca = ["CA"]
    src_ca  = ["CAL"]

    # 4) Canonical data fields
    data_fields = [
        "Service Revenue","Proppant Revenue","Proppant Handling Revenue","Chemical Revenue","Fuel Revenue",
        "Total Revenue","Proppant Costs","Proppant Logistic Cost","Chemical and Gel","Mat and Containment Costs",
        "Fuel Costs","Other Pad Costs","Allocation - VM","Total Variable Cost","Variable Margin","Variable Margin %",
        "RM - Parts and Services","RM - Fluid End/Pumps","RM - IRON","RM - Startup","RM - Interbasin Reclass",
        "RM - Mining Adjustment","RM - Allocation","RM - Manual Allocation","R&M & Reclass (Net)",
        "Personnel - Salaries and Benefit","Personnel - Contractors","Personnel - Overtime","Personnel - Job Bonus",
        "Personnel - Quarterly Bonus","Personnel - Stock Comp","Personnel - Mancamp","Personnel - Other",
        "Personnel - Startup","Personnel - Interbasin Reclass","Personnel - Mining Adjustment","Personnel - Allocation",
        "Personnel Cost & Reclass (Net)","District - Equipment Rental","District - Sand Box Rental","District - Property Tax",
        "District - Other District Cost","District - Interbasin Reclass & Startup","District - Mining Adjustment",
        "District - Allocation","District Costs & Reclass (Net)","Cost of Sales","Gross Margin","Gross Margin %",
        "G&A - Stock Comp","G&A - Personnel","G&A - Non-Stock Comp and Personnel","G&A - Allocation",
        "G&A - Manual Allocation","Total G&A","Corporate - Manual Allocation","EBITDA","EBITDA%","DD&A","Interest & Fees",
        "Non Recurring","Startup Expense","Mining Adjustment - DD&A","Corporate Allocation","Corporate Allocation-Manual DD&A",
        "IBT","IBT%","Staffed Fleet","Utilized Fleet (Standard Days Home Crew)","Stages","Pump Days",
        "Standard Days (Work Basin)","Standard Days (Home Crew)","Pump Hrs","HHP Hr (K)",
        "Proppant Pumped","Proppant Sold","Proppant Hauled","Headcount (Employees)",
        "Headcount (Contractors)","Total Headcount","Calandar Day"
    ]

    # 5) Field‐name aliases for variations
    field_aliases = {
        "District - Other District Cost": ["Other District Cost"],
    }

    # 6) Locate the target‐month column in row 3
    def find_month_col(ws):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=3, column=col).value
            if isinstance(val, date):
                cand = val.strftime("%b-%y")
            elif isinstance(val, str):
                cand = val.strip()
            else:
                continue
            if cand == target_month_str:
                return col
        return None

    col_us = find_month_col(ws_us)
    col_ca = find_month_col(ws_ca)
    if not col_us or not col_ca:
        logging.error(f"Month '{target_month_str}' not found in row 3 of US/CA.")
        sys.exit(f"Cannot find column for {target_month_str}")

    # 7) Compute date columns
    dt        = datetime.strptime(target_month_str, "%b-%y")
    year      = dt.year
    month     = dt.month
    last_day  = calendar.monthrange(year, month)[1]
    report_date = date(year, month, last_day)
    month_name  = report_date.strftime("%B")
    quarter     = (month - 1) // 3 + 1
    m_y         = target_month_str
    q_y         = f"Q{quarter}-{str(year)[-2:]}"

    # 8) Build in‐memory pivot table
    headers = ["Basin","Date","Year","Month","Q","M-Y","Q-Y"] + data_fields
    pivot   = [headers]

    def extract(ws, dest_list, src_list, month_col):
        maxr = ws.max_row
        for idx, basin_dest in enumerate(dest_list):
            basin_src = src_list[idx]
            # find basin row in col A
            basin_row = next(
                (r for r in range(1, maxr+1)
                 if str(ws.cell(r,1).value or "").strip() == basin_src),
                None
            )
            if not basin_row:
                logging.warning(f"{basin_src} not found in {ws.title}")
                continue

            row_data = [
                basin_dest,
                report_date.isoformat(),
                year,
                month_name,
                quarter,
                m_y,
                q_y
            ]

            for df in data_fields:
                val = "N/A"
                variants = [df] + field_aliases.get(df, [])
                for r2 in range(basin_row, maxr+1):
                    nxt = ws.cell(r2,1).value
                    if r2>basin_row and nxt and str(nxt).strip() in src_list and str(nxt).strip()!= basin_src:
                        break
                    cell_f = str(ws.cell(r2,6).value or "").strip()
                    if cell_f in variants:
                        val = ws.cell(r2, month_col).value
                        break
                row_data.append(val)

            pivot.append(row_data)

    # Extract US & CA
    extract(ws_us, dest_us, src_us, col_us)
    extract(ws_ca, dest_ca, src_ca, col_ca)

    # --- NEW: process AU sheet exactly the same way ---
    au_sheet = "NIS Details by Basin - AU"
    if au_sheet in wb_fin.sheetnames:
        ws_au = wb_fin[au_sheet]
        col_au = find_month_col(ws_au)
        if not col_au:
            logging.warning(f"Month '{target_month_str}' not in row 3 of '{au_sheet}'. Skipping AU.")
        else:
            # fill down blanks in col A
            last = None
            for r in range(1, ws_au.max_row+1):
                v = ws_au.cell(r,1).value
                if v:
                    last = str(v).strip()
                else:
                    ws_au.cell(r,1).value = last

            dest_au = ["AU"]
            src_au  = ["AU"]
            extract(ws_au, dest_au, src_au, col_au)
    else:
        logging.info(f"'{au_sheet}' not found; skipping AU extraction.")

    # 9) Append into the reference workbook
    try:
        wb_ref = load_workbook(ref_file)
    except FileNotFoundError:
        wb_ref = Workbook()
        wb_ref.remove(wb_ref.active)

    sheet_name = "CK data Pivot"
    if sheet_name in wb_ref.sheetnames:
        ws_ref = wb_ref[sheet_name]
        start  = ws_ref.max_row + 1
    else:
        ws_ref = wb_ref.create_sheet(sheet_name)
        # write header row
        for j, h in enumerate(pivot[0], start=1):
            ws_ref.cell(row=1, column=j, value=h)
        start = 2

    # write data rows
    for i, row in enumerate(pivot[1:], start=0):
        for j, v in enumerate(row, start=1):
            ws_ref.cell(row=start + i, column=j, value=v)

    wb_ref.save(ref_file)
    logging.info(f"Appended {len(pivot)-1} rows to '{sheet_name}' in {ref_file}.")

# modules/report_generation.py

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from settings_manager import load_config

def build_monthly_database(
    tableau_exports_path: str,
    output_dir: str,
    config=None,
) -> str:
    """
    Create a 'Database' sheet in a fresh workbook saved into
    `output_dir/Monthly_Database.xlsx` that
    lays out, side-by-side, each source table from:
      - the Tableau export workbook (tableau_exports_path)
      - the RefData workbook configured in config.ini

    Each block is headed by a merged cell with the view name, then the
    column headers on row 2, and data beginning on row 3.  One blank column
    is left between blocks.
    """
    # === define which sheets to pull from where ===
    tableau_views = [
        "Main_Combo",
        "Stragglers",
        "Basin_Crew_Count",
        "GL_Basin_Pivot",
        "Project_VM",
        "Transload",
        "Unassigned_Rev_CA",
        "Unalloc_Costs",
        "FCAST",
        "Project_List",
    ]
    ref_views = [
        "FX",
        "CK data Pivot"
    ]

    if config is None:
        config = load_config()
    ref_data_path = config["files"].get("ref_data_path", "")
    output_path = os.path.join(output_dir, "Monthly_Database.xlsx")

    # load workbooks via pandas
    tbl_xl = pd.ExcelFile(tableau_exports_path)
    ref_xl = pd.ExcelFile(ref_data_path)

    # start fresh
    wb = Workbook()
    ws = wb.active
    ws.title = "Database"

    current_col = 1  # 1-based Excel column

    def write_block(df: pd.DataFrame, title: str, start_col: int):
        """Helper: write one DataFrame block into ws, return next column to use."""
        n_cols = df.shape[1]
        end_col = start_col + n_cols - 1

        # 1) merged title on row 1
        ws.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1,   end_column=end_col
        )
        ws.cell(row=1, column=start_col, value=title)

        # 2) column headers on row 2
        for i, col_name in enumerate(df.columns, start=start_col):
            ws.cell(row=2, column=i, value=col_name)

        # 3) data starting on row 3
        for r_idx, row in enumerate(df.values, start=3):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 4) return next free column (one blank column)
        return end_col + 2

    # 1) pull each Tableau view
    for view in tableau_views:
        if view in tbl_xl.sheet_names:
            df = tbl_xl.parse(view)
            current_col = write_block(df, view, current_col)
        else:
            # skip if user removed this view
            continue

    # 2) pull each RefData view
    for view in ref_views:
        if view in ref_xl.sheet_names:
            df = ref_xl.parse(view)
            current_col = write_block(df, view, current_col)
        else:
            continue

    # 3) save
    wb.save(output_path)
    print(f"Built monthly database sheet -> {output_path}")
    return output_path


if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Usage: python report_generation.py <tableau_exports.xlsx> <output_dir>")
        sys.exit(1)
    build_monthly_database(sys.argv[1], sys.argv[2])

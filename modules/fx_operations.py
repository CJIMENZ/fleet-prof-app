"""
modules/fx_operations.py

Handles:
  - create_new_workbook(oracle_usd_path, oracle_cad_path) for CAD
  - calculate_fx_n5(...) for CAD→USD
  - calculate_fx_n5_aud(...) for AUD→USD
  - clean_cad_data(...)
  - update_fx_ref(...) helper
"""

import os
import sys
import logging
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import messagebox

def create_new_workbook(oracle_usd_path, oracle_cad_path):
    """
    Copies 'LOS Management Report IS19' from the Oracle USD and CAD files
    into a new workbook with sheets "USD" and "CA".
    (If you also need a separate workbook for AUD, you'd define a similar function.)
    """
    logging.info("Creating new workbook with USD and CA.")
    try:
        usd_wb = load_workbook(oracle_usd_path)
        cad_wb = load_workbook(oracle_cad_path)

        usd_sheet = usd_wb["LOS Management Report IS19"]
        cad_sheet = cad_wb["LOS Management Report IS19"]

        new_wb = Workbook()
        default_sheet = new_wb.active
        new_wb.remove(default_sheet)

        # Copy USD
        new_wb.create_sheet("USD")
        target_usd = new_wb["USD"]
        for row in usd_sheet.iter_rows(values_only=True):
            target_usd.append(row)

        # Copy CA
        new_wb.create_sheet("CA")
        target_ca = new_wb["CA"]
        for row in cad_sheet.iter_rows(values_only=True):
            target_ca.append(row)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder = os.path.dirname(oracle_usd_path)
        new_file = os.path.join(folder, f"1.Oracle_Variance_CAD_{timestamp}.xlsx")
        new_wb.save(new_file)
        logging.info(f"New workbook created: {new_file}")
        return new_file
    except Exception as e:
        logging.error(f"Error in create_new_workbook: {e}")
        sys.exit("Failed to create new workbook.")

def calculate_fx_n5(new_workbook_path, ref_table_path):
    """
    Reads cell N5 in 'USD' & 'CA' => ratio = CA!N5 / USD!N5.
    Identifies the last column header (row=2) in 'USD' => latest_month.
    Calls update_fx_ref(..., fx_col_name='CAD/USD').
    Returns (fx, latest_month).
    """
    logging.info("Calculating CAD→USD FX (N5).")
    try:
        wb = load_workbook(new_workbook_path)
        usd_sheet = wb["USD"]
        ca_sheet  = wb["CA"]

        usd_val = usd_sheet["N5"].value
        ca_val  = ca_sheet["N5"].value
        if not usd_val or not ca_val:
            raise ValueError("N5 is empty or invalid in USD/CA sheets (CAD).")

        fx = ca_val / usd_val
        logging.info(f"N5 => CA={ca_val}, USD={usd_val}, ratio={fx}")

        max_col = usd_sheet.max_column
        last_header = usd_sheet.cell(row=2, column=max_col).value
        latest_month = str(last_header) if last_header else "UnknownMonth"

        logging.info(f"latest_month={latest_month}, col={get_column_letter(max_col)}")

        update_fx_ref(ref_table_path, latest_month, fx, fx_col_name="CAD/USD")
        wb.save(new_workbook_path)
        logging.info("Saved workbook after CAD→USD FX calc.")
        return fx, latest_month
    except Exception as e:
        logging.error(f"Error in calculate_fx_n5: {e}")
        sys.exit("CAD→USD calculation failed.")

def calculate_fx_n5_aud(oracle_usd_path, oracle_aud_path, ref_table_path, existing_month=None):
    """
    Scans column N (from row 5 downward) in:
      - 'LOS Management Report IS19' of the USD file
      - 'LOS Management Report IS29' of the AUD file
    Finds the first non-zero pair => ratio = (AUD value) / (USD value).
    Then appends "AUD/USD" to the RefData 'FX' sheet (for the same month as existing_month
    if provided, else we identify the last column header from the USD sheet).
    """
    logging.info("Calculating AUD→USD FX by scanning column N for the first non-zero row.")
    try:
        # 1) Load the Oracle USD workbook, sheet "LOS Management Report IS19"
        usd_wb = load_workbook(oracle_usd_path, data_only=True)
        usd_sheet = usd_wb["LOS Management Report IS29"]

        # 2) Load the Oracle AUD workbook, sheet "LOS Management Report IS29"
        aud_wb = load_workbook(oracle_aud_path, data_only=True)
        aud_sheet = aud_wb["LOS Management Report IS29"]

        # 3) Scan column N (i.e., column index=14 if 1-based) from row 5 downward
        usd_ratio_val = None
        aud_ratio_val = None
        max_row_usd = usd_sheet.max_row
        max_row_aud = aud_sheet.max_row
        found_ratio = False

        # We'll iterate up to the min of (max_row_usd, max_row_aud) in column N
        for row_idx in range(5, min(max_row_usd, max_row_aud) + 1):
            usd_val = usd_sheet.cell(row=row_idx, column=14).value  # column N is col index=14
            aud_val = aud_sheet.cell(row=row_idx, column=14).value
            if usd_val and aud_val:
                try:
                    usd_num = float(usd_val)
                    aud_num = float(aud_val)
                    if usd_num != 0 and aud_num != 0:
                        usd_ratio_val = usd_num
                        aud_ratio_val = aud_num
                        found_ratio = True
                        logging.info(f"Found non-zero pair at row {row_idx}: AUD={aud_num}, USD={usd_num}")
                        break
                except:
                    # skip if can't convert
                    continue

        if not found_ratio:
            raise ValueError("No valid non-zero pair found in column N for AUD→USD calculation.")

        ratio = aud_ratio_val / usd_ratio_val
        logging.info(f"AUD→USD ratio found: AUD={aud_ratio_val}, USD={usd_ratio_val}, ratio={ratio}")

        # 4) Determine latest_month
        if existing_month:
            latest_month = existing_month
            logging.info(f"Using existing month: {existing_month} for AUD→USD")
        else:
            # If none provided, read the last column header in row=2 of the USD sheet
            max_col = usd_sheet.max_column
            last_header = usd_sheet.cell(row=2, column=max_col).value
            latest_month = str(last_header) if last_header else "UnknownMonth"
            logging.info(f"Derived latest_month={latest_month} from USD last column header.")

        # 5) Append AUD/USD to the RefData 'FX' sheet
        update_fx_ref(ref_table_path, latest_month, ratio, fx_col_name="AUD/USD")

    except Exception as e:
        logging.error(f"Error in calculate_fx_n5_aud: {e}")
        sys.exit("AUD→USD calculation failed.")


def update_fx_ref(ref_file_path, latest_month, fx, fx_col_name="CAD/USD"):
    """
    Overwrites/appends row in 'FX' sheet => columns [Date, fx_col_name].
    For CAD, fx_col_name='CAD/USD', for AUD, fx_col_name='AUD/USD'.
    """
    from openpyxl import load_workbook
    import pandas as pd

    logging.info(f"Updating FX in reference workbook (col={fx_col_name}).")
    try:
        # Ensure the ref file
        try:
            wb = load_workbook(ref_file_path)
        except FileNotFoundError:
            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            wb.save(ref_file_path)

        # Read existing "FX" with pandas
        try:
            fx_df = pd.read_excel(ref_file_path, sheet_name="FX")
        except:
            fx_df = pd.DataFrame()

        if fx_df.empty:
            fx_df = pd.DataFrame(columns=["Date", fx_col_name])
        else:
            # If 'Period' => rename to 'Date'
            if "Date" not in fx_df.columns and "Period" in fx_df.columns:
                fx_df.rename(columns={"Period": "Date"}, inplace=True)

            if "Date" not in fx_df.columns:
                fx_df["Date"] = None
            if fx_col_name not in fx_df.columns:
                fx_df[fx_col_name] = None

        # Overwrite or append
        str_dates = fx_df["Date"].astype(str).values
        if str(latest_month) in str_dates:
            mask = fx_df["Date"].astype(str) == str(latest_month)
            fx_df.loc[mask, fx_col_name] = float(fx)
            logging.info(f"Overwrote {fx_col_name} for {latest_month} => {fx}")
        else:
            new_row = {"Date": latest_month, fx_col_name: float(fx)}
            fx_df = pd.concat([fx_df, pd.DataFrame([new_row])], ignore_index=True)
            logging.info(f"Appended new FX => [Date={latest_month}, {fx_col_name}={fx}]")

        # Create a new workbook with the FX sheet
        new_wb = Workbook()
        default_sheet = new_wb.active
        new_wb.remove(default_sheet)
        
        # Create the FX sheet
        fx_sheet = new_wb.create_sheet("FX")
        
        # Write the data
        for r_idx, row in enumerate(dataframe_to_rows(fx_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                fx_sheet.cell(row=r_idx, column=c_idx, value=value)

        # Save the workbook
        new_wb.save(ref_file_path)

    except Exception as e:
        logging.error(f"Error in update_fx_ref: {e}")
        sys.exit("Failed to update FX table.")


def clean_cad_data(new_workbook_path, latest_month):
    """
    Existing logic for cleaning CAD data => 'Data Sort CAD'.
    """
    logging.info("Cleaning CAD data.")
    try:
        wb = load_workbook(new_workbook_path)
        ca_sheet = wb["CA"]

        headers = [cell.value for cell in ca_sheet[2]]
        if latest_month not in headers:
            logging.warning(f"Month {latest_month} not in CA headers.")
            use_latest = messagebox.askyesno(
                "Month Mismatch",
                f"'{latest_month}' not in CA headers.\nUse the last header instead?"
            )
            if use_latest and headers:
                latest_month = headers[-1]
                latest_col = len(headers)
            else:
                sys.exit("No valid month found in CA headers.")
        else:
            latest_col = headers.index(latest_month) + 1

        data_sort_cad = []
        sn = 1
        for row_idx in range(5, ca_sheet.max_row + 1):
            acct_val = ca_sheet.cell(row=row_idx, column=1).value
            if acct_val:
                parts = str(acct_val).split('-')
                acct_num = parts[0].strip() if parts else ""
                if acct_num.isdigit() and len(acct_num) == 6:
                    desc = ca_sheet.cell(row=row_idx, column=2).value
                    val  = ca_sheet.cell(row=row_idx, column=latest_col).value
                    data_sort_cad.append({
                        "SN": sn,
                        "REVENUE": desc,
                        "Account2": acct_num,
                        latest_month: val
                    })
                    sn += 1

        if not data_sort_cad:
            logging.error("No valid 6-digit accounts in CA sheet.")
            sys.exit("No valid 6-digit accounts in CA.")

        df_cad = pd.DataFrame(data_sort_cad)

        if "Data Sort CAD" in wb.sheetnames:
            del wb["Data Sort CAD"]
        ds_sheet = wb.create_sheet("Data Sort CAD")

        for r in dataframe_to_rows(df_cad, index=False, header=True):
            ds_sheet.append(r)

        wb.save(new_workbook_path)
        logging.info("Data Sort CAD created.")
        return new_workbook_path, latest_month
    except Exception as e:
        logging.error(f"Error cleaning CAD data: {e}")
        sys.exit("Failed to clean CAD data.")

# modules/pnl_pivot_operations.py
import logging
from typing import List, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# --------------------------------------------------------------------------- #
#   CONSTANTS / COLUMN MAPS
# --------------------------------------------------------------------------- #
# CK block in Database sheet            DW (127) … HG (208)
_CK_FIRST_COL = "DW"
_CK_LAST_COL  = "HG"

# Project‑VM block                      AO (41) … BD (56)
_VM_FIRST_NUM = 41
_VM_LAST_NUM  = 56

# FX block                              123–125
_FX_COL_DATE   = 123   # FX – Date
_FX_COL_CADUSD = 124   # FX – CAD/USD

# ---------- Basin aliases ---------------------------------------------------
# Any spelling on the LEFT will be treated as the canonical label on the RIGHT
_BASIN_ALIAS: Dict[str, str] = {
    "CORP":       "Corporate",
    "CORPORATE":  "Corporate",
    "CORP.":      "Corporate",
}

def _canonical(basin: str | None) -> str:
    """Return canonical basin name for comparison matching."""
    if basin is None:
        return ""
    key = str(basin).strip().upper()
    return _BASIN_ALIAS.get(key, str(basin).strip())

# ---------- table field definitions ----------------------------------------
_PNL_FIELDS: List[str] = [
    "Total Revenue", "Total Variable Cost",
    "R&M & Reclass (Net)", "Personnel Cost & Reclass (Net)",
    "District Costs & Reclass (Net)", "Gross Margin",
    "Total G&A", "Corporate - Manual Allocation", "EBITDA",
]

_CK_FIELDS: List[str] = [
    "SERVICE REV", "PROP REV", "TRUCK REV", "CHEM REV", "FUEL REV",
    "PROP COST", "TRUCK COST", "CHEM COST", "FUEL COST",
    "MAT COST", "OTHER PAD COST", "ALLOC VM COST"
]
_CK_COL_NUM = {
    # revenues
    "SERVICE REV": 134, "PROP REV": 135, "TRUCK REV": 136,
    "CHEM REV":    137, "FUEL REV": 138,
    # costs
    "PROP COST": 140, "TRUCK COST": 141, "CHEM COST": 142,
    "FUEL COST": 144, "MAT COST": 143, "OTHER PAD COST": 145,
    "ALLOC VM COST": 146,
    # keys
    "Basin": 127, "M-Y": 132,
}

_VM_FIELDS = _CK_FIELDS.copy()
_VM_COL_NUM = {
    "SERVICE REV": 45, "PROP REV": 46, "TRUCK REV": 47, "CHEM REV": 48,
    "FUEL REV": 49,
    "PROP COST": 50, "TRUCK COST": 51, "CHEM COST": 52, "FUEL COST": 53,
    "MAT COST": 54, "OTHER PAD COST": 55, "ALLOC VM COST": 56,
    "Basin": 43,
}

_REVENUE_FIELDS = {"SERVICE REV", "PROP REV", "TRUCK REV",
                   "CHEM REV", "FUEL REV"}

COL_WIDTH = 25   # spreadsheet aesthetics only

# --------------------------------------------------------------------------- #
def generate_pnl_pivot(month_data_path: str) -> None:
    logging.info(f"Building all pivot tables in {month_data_path}")
    wb   = load_workbook(month_data_path)
    wsdb = wb["Database"]

    # ---------------- collect CK block into DataFrame -----------------------
    start_ck = wsdb[f"{_CK_FIRST_COL}2"].column
    end_ck   = wsdb[f"{_CK_LAST_COL}2"].column
    ck_headers = list(next(wsdb.iter_rows(
        min_row=2, max_row=2, min_col=start_ck, max_col=end_ck, values_only=True
    )))
    ck_rows = list(wsdb.iter_rows(
        min_row=3, min_col=start_ck, max_col=end_ck, values_only=True
    ))
    df_ck = pd.DataFrame(ck_rows, columns=ck_headers)

    # ---------------- distinct basin names ----------------------------------
    basins_ck = sorted({b for b in df_ck["Basin"].dropna().astype(str)})
    vm_basin_vals = (
        wsdb.cell(row=r, column=_VM_COL_NUM["Basin"]).value
        for r in range(3, wsdb.max_row + 1)
    )
    basins_vm = sorted({
        str(b).strip() for b in vm_basin_vals
        if b not in (None, "", " ")
    })

    months  = sorted(df_ck["M-Y"].dropna().astype(str).unique())
    latest_month = months[-1] if months else ""
    last_db_row  = wsdb.max_row

    # ---------------- FX (CAD/USD) lookup -----------------------------------
    fx_rate = _find_cad_usd_rate(wsdb, latest_month)
    logging.info(f"CAD/USD for {latest_month} -> {fx_rate}")

    # ---------------- rebuild PnL Pivot sheet --------------------------------
    if "PnL Pivot" in wb.sheetnames:
        del wb["PnL Pivot"]
    ws = wb.create_sheet("PnL Pivot")

    # ---- dropdown for month -------------------------------------------------
    ws["A1"] = "Select month in B2 – CK tables auto‑recalc."
    ws["A2"] = "M-Y"
    dv = DataValidation(type="list", formula1=f'"{",".join(months)}"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["B2"])
    ws["B2"].value = latest_month
    ws["B2"].fill = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")

    # ---- helper cell for CAD/USD -------------------------------------------
    ws["Z2"] = fx_rate
    ws.column_dimensions["Z"].hidden = True

    bold = Font(bold=True)
    hdr_align = Alignment(horizontal="center", wrap_text=True)

    row_ptr = 4  # rolling pointer down the sheet

    # 1) CK PnL ---------------------------------------------------------------
    row_ptr = _write_block(
        ws, "CK PnL", row_ptr, basins_ck, _PNL_FIELDS,
        _header_to_letter_map(ck_headers, start_ck),
        basin_col=get_column_letter(_CK_COL_NUM["Basin"]),
        my_col=get_column_letter(_CK_COL_NUM["M-Y"]),
        last_db_row=last_db_row,
        use_month=True,
        bold=bold, hdr_align=hdr_align
    )

    # 2) CK VM ----------------------------------------------------------------
    row_ptr = _write_block(
        ws, "CK VM", row_ptr + 2, basins_ck, _CK_FIELDS,
        {f: get_column_letter(_CK_COL_NUM[f]) for f in _CK_FIELDS},
        basin_col=get_column_letter(_CK_COL_NUM["Basin"]),
        my_col=get_column_letter(_CK_COL_NUM["M-Y"]),
        last_db_row=last_db_row,
        use_month=True,
        bold=bold, hdr_align=hdr_align
    )
    ck_hdr_row  = row_ptr - len(basins_ck) - 2   # header row for CK VM
    ck_data_row0 = ck_hdr_row + 1
    ck_data_last = ck_data_row0 + len(basins_ck) - 1

    # 3) Project VM (CAD->USD) -------------------------------------------------
    row_ptr = _write_block(
        ws, "Project VM", row_ptr + 2, basins_vm, _VM_FIELDS,
        {f: get_column_letter(_VM_COL_NUM[f]) for f in _VM_FIELDS},
        basin_col=get_column_letter(_VM_COL_NUM["Basin"]),
        my_col="",            # unused (no month filter)
        last_db_row=last_db_row,
        use_month=False,
        cad_fx_cell="$Z$2",
        bold=bold, hdr_align=hdr_align
    )
    vm_hdr_row  = row_ptr - len(basins_vm) - 2   # header row for P‑VM
    vm_data_row0 = vm_hdr_row + 1
    vm_data_last = vm_data_row0 + len(basins_vm) - 1

    # ------ build row maps (canonical name -> worksheet row) -----------------
    ck_row_map = {_canonical(b): ck_data_row0 + i
                  for i, b in enumerate(basins_ck)}
    vm_row_map = {_canonical(b): vm_data_row0 + i
                  for i, b in enumerate(basins_vm)}

    # 4) Comparison -----------------------------------------------------------
    comparison_basins = sorted(set(ck_row_map) | set(vm_row_map))
    row_ptr = _write_comparison_block(
        ws, "Comparison", row_ptr + 2,
        comparison_basins, _CK_FIELDS,
        ck_row_map=ck_row_map,
        vm_row_map=vm_row_map,
        bold=bold, hdr_align=hdr_align
    )

    wb.save(month_data_path)
    logging.info("PnL Pivot sheet rebuilt successfully.")

# --------------------------------------------------------------------------- #
#  Helper: find CAD/USD for the selected month
def _find_cad_usd_rate(wsdb, month_label: str) -> float:
    """Return CAD/USD matching month_label, else last numeric in col 124."""
    from datetime import datetime
    for r in range(3, wsdb.max_row + 1):
        dt = wsdb.cell(r, _FX_COL_DATE).value
        cand = dt.strftime("%b-%y") if isinstance(dt, datetime) else str(dt).strip()
        if cand == month_label:
            try:
                return float(wsdb.cell(r, _FX_COL_CADUSD).value)
            except (TypeError, ValueError):
                break
    for r in range(wsdb.max_row, 2, -1):
        try:
            return float(wsdb.cell(r, _FX_COL_CADUSD).value)
        except (TypeError, ValueError):
            continue
    return 1.0  # should not happen

# --------------------------------------------------------------------------- #
#  Write a summary block and return the next free row
def _write_block(
    ws, title: str, start_row: int,
    basins: List[str], fields: List[str], col_map: Dict[str, str],
    basin_col: str, my_col: str, last_db_row: int, use_month: bool,
    bold: Font, hdr_align: Alignment, cad_fx_cell: str = None
) -> int:
    """Generic writer for CK PnL, CK VM, Project VM blocks."""
    hdr_row = start_row + 1
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row,   end_column=len(fields) + 1)
    ws.cell(start_row, 1, title).font = bold
    ws.cell(start_row, 1).alignment = Alignment(horizontal="center")

    ws.row_dimensions[hdr_row].height = 30
    for col, head in enumerate(["Basin", *fields], start=1):
        cell = ws.cell(hdr_row, col, head)
        cell.font = bold
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH

    # data rows
    for r_off, basin in enumerate(basins, start=1):
        r = hdr_row + r_off
        ws.cell(r, 1, basin).alignment = Alignment(horizontal="center")
        for f_off, field in enumerate(fields, start=1):
            data_col = col_map[field]
            sum_range = f"${data_col}$3:${data_col}${last_db_row}"
            basin_rng = f"${basin_col}$3:${basin_col}${last_db_row}"
            if use_month:
                my_rng = f"${my_col}$3:${my_col}${last_db_row}"
                base = (f"SUMIFS(Database!{sum_range},"
                        f"Database!{my_rng},$B$2,"
                        f"Database!{basin_rng},$A{r})")
            else:
                base = (f"SUMIF(Database!{basin_rng},$A{r},"
                        f"Database!{sum_range})")
            formula = (f"=IF($A{r}=\"CA\",({base})/{cad_fx_cell},({base}))"
                       if cad_fx_cell else f"={base}")
            cell = ws.cell(r, f_off + 1, formula)
            cell.number_format = '$#,##0.00'

    grand = hdr_row + len(basins) + 1
    ws.cell(grand, 1, "Grand Total").font = bold
    ws.cell(grand, 1).alignment = Alignment(horizontal="center")
    for c in range(2, len(fields) + 2):
        L = get_column_letter(c)
        ws.cell(grand, c,
                f"=SUM({L}{hdr_row+1}:{L}{grand-1})").number_format = '$#,##0.00'
        ws.cell(grand, c).font = bold

    return grand + 1

# --------------------------------------------------------------------------- #
def _write_comparison_block(
    ws, title: str, start_row: int, basins: List[str], fields: List[str],
    ck_row_map: Dict[str, int], vm_row_map: Dict[str, int],
    bold: Font, hdr_align: Alignment
) -> int:
    """CK – VM (rev) or CK + VM (cost) by canonical basin name."""
    hdr_row = start_row + 1
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row,   end_column=len(fields) + 1)
    ws.cell(start_row, 1, title).font = bold
    ws.cell(start_row, 1).alignment = Alignment(horizontal="center")

    # header row
    ws.row_dimensions[hdr_row].height = 30
    for col, head in enumerate(["Basin", *fields], start=1):
        cell = ws.cell(hdr_row, col, head)
        cell.font = bold
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH

    # data rows
    for i, canon in enumerate(basins):
        r = hdr_row + 1 + i
        ws.cell(r, 1, canon).alignment = Alignment(horizontal="center")

        ck_row = ck_row_map.get(canon)
        vm_row = vm_row_map.get(canon)

        for j, field in enumerate(fields, start=1):
            col_letter = get_column_letter(j + 1)   # same col layout in both blocks
            ck_cell = f"{col_letter}{ck_row}" if ck_row else "0"
            vm_cell = f"{col_letter}{vm_row}" if vm_row else "0"
            if field in _REVENUE_FIELDS:
                formula = f"={ck_cell}-{vm_cell}"
            else:
                formula = f"={ck_cell}+{vm_cell}"
            cell = ws.cell(r, j + 1, formula)
            cell.number_format = '$#,##0.00'

    # grand total
    grand = hdr_row + len(basins) + 1
    ws.cell(grand, 1, "Grand Total").font = bold
    ws.cell(grand, 1).alignment = Alignment(horizontal="center")
    for c in range(2, len(fields) + 2):
        L = get_column_letter(c)
        ws.cell(grand, c,
                f"=SUM({L}{hdr_row+1}:{L}{grand-1})").number_format = '$#,##0.00'
        ws.cell(grand, c).font = bold

    return grand + 1

# --------------------------------------------------------------------------- #
def _header_to_letter_map(headers: List[str], start_col_num: int) -> Dict[str, str]:
    return {h: get_column_letter(idx)
            for h, idx in zip(headers, range(start_col_num,
                                             start_col_num + len(headers)))}

# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO)
    if len(sys.argv) != 2:
        sys.exit("Usage: python pnl_pivot_operations.py <MonthDataFile.xlsx>")
    generate_pnl_pivot(sys.argv[1])

#!/usr/bin/env python3
"""
Finance‑grade Project VM workbook builder
----------------------------------------

Creates (in this order):
    1. P. VM ‑ All USD
    2. P. VM ‑ Current
    3. P. VM ‑ Previous
    4. P. VM ‑ Unalloc
    5. P. VM ‑ Unass
    6. P. VM ‑ Adjustments   (summary, split buckets, manual grid)

Key rules
~~~~~~~~~
* CAD -> USD rate = last numeric value in column 124
* AUD -> USD rate = last numeric value in column 125
* Project_VM numeric fields pulled by **fixed indices 45‑56** — duplicate
  headers elsewhere can't interfere.
* All headers are stripped (`.strip()`) immediately after load.
* Every detail sheet gets a bold Grand Total row inserted as row 2.

Run:
    python main.py <MonthlyDataFile.xlsx>
"""

# -------------------------------------------------------------------- #
#  Imports & logging
# -------------------------------------------------------------------- #
import sys, datetime, logging, zipfile, shutil, re
from typing import List, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

logging.basicConfig(level=logging.INFO,
                    format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# -------------------------------------------------------------------- #
#  CONSTANTS
# -------------------------------------------------------------------- #
COL_PROJ_NUM, COL_ENG_BASIN = 42, 43
COL_MAIN_PAD, COL_PREV_PAD  = 1, 115
COL_FX_CAD,  COL_FX_AUD     = 124, 125

VM_COL_IDX = {
    "Service Rev":                  45,
    "Prop Rev":                     46,
    "Proppant Handling Revenue":    47,
    "Chem Rev":                     48,
    "Fuel Rev":                     49,
    "Prop Cost":                    50,
    "Truck Cost":                   51,
    "Chemical and Gel cost":        52,
    "Fuel Cost":                    53,
    "Mat and Containment Costs":    54,
    "Other Pad Costs":              55,
    "Allocation VM":                56,
}

DB_SHEET        = "Database"
HEADER_ROW      = 2
PIVOT_MARKER    = "Q-Y"
PIVOT_MONTH_COL = "M-Y"

# cost fields to flip sign in CK vs VM comparison
COST_FIELDS = {
    "PROP COST","TRUCK COST","CHEM COST","FUEL COST",
    "MAT COST","OTHER PAD COST","ALLOC VM COST"
}

# -------------------------------------------------------------------- #
#  CK & VM mappings  (unchanged from previous version)
# -------------------------------------------------------------------- #
CK_COLUMNS: Dict[str, str] = {
    "REVENUE":        "Total Revenue",
    "SERVICE REV":    "Service Revenue",
    "PROP REV":       "Proppant Revenue",
    "TRUCK REV":      "Proppant Handling Revenue",
    "CHEM REV":       "Chemical Revenue",
    "FUEL REV":       "Fuel Revenue",
    "MISC REV":       None,
    "VARIABLE COST":  "Total Variable Cost",
    "PROP COST":      "Proppant Costs",
    "TRUCK COST":     "Proppant Logistic Cost",
    "CHEM COST":      "Chemical and Gel",
    "FUEL COST":      "Fuel Costs",
    "MAT COST":       "Mat and Containment Costs",
    "OTHER PAD COST": "Other Pad Costs",
    "ALLOC VM COST":  "Allocation - VM",
    "MISC COST":      None,
}

VM_SUM_MAP: Dict[str, List[str]] = {
    "REVENUE":        ["Service Rev","Prop Rev","Proppant Handling Revenue",
                       "Chem Rev","Fuel Rev"],
    "SERVICE REV":    ["Service Rev"],
    "PROP REV":       ["Prop Rev"],
    "TRUCK REV":      ["Proppant Handling Revenue"],
    "CHEM REV":       ["Chem Rev"],
    "FUEL REV":       ["Fuel Rev"],
    "MISC REV":       [],
    "VARIABLE COST":  ["Prop Cost","Truck Cost","Chemical and Gel cost",
                       "Fuel Cost","Mat and Containment Costs",
                       "Other Pad Costs","Allocation VM"],
    "PROP COST":      ["Prop Cost"],
    "TRUCK COST":     ["Truck Cost"],
    "CHEM COST":      ["Chemical and Gel cost"],
    "FUEL COST":      ["Fuel Cost"],
    "MAT COST":       ["Mat and Containment Costs"],
    "OTHER PAD COST": ["Other Pad Costs"],
    "ALLOC VM COST":  ["Allocation VM"],
    "MISC COST":      [],
}
SUMMARY_FIELDS = list(VM_SUM_MAP.keys())

_FLAG_COLS = ["COMMENT"]
MANUAL_COLUMNS = ["Project Number","ENG BASIN R1","Period Name"] \
                 + SUMMARY_FIELDS + _FLAG_COLS

# -------------------------------------------------------------------- #
#  Helper utilities (unchanged)
# -------------------------------------------------------------------- #
_MONEY_RE = re.compile(r"[,$]")

def _to_number(series: pd.Series) -> pd.Series:
    cleaned = (series.astype(str)
                     .str.replace(_MONEY_RE, "", regex=True)
                     .str.replace(r"\((.*)\)", r"-\1", regex=True))
    return pd.to_numeric(cleaned, errors="coerce")

def _safe_load_workbook(path: str, backup: bool = False):
    with zipfile.ZipFile(path, "r") as zf:
        zf.testzip()
    if backup:
        shutil.copy2(path, path.replace(".xlsx", "-backup.xlsx"))
    return load_workbook(path)

def _read_block(ws, first: str, last: str) -> pd.DataFrame:
    s = ws[f"{first}{HEADER_ROW}"].column
    e = ws[f"{last}{HEADER_ROW}"].column
    hdr = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW,
                            min_col=s, max_col=e, values_only=True))
    cols = [h.strip() if isinstance(h, str) else h for h in hdr]
    data = list(ws.iter_rows(min_row=HEADER_ROW+1, min_col=s, max_col=e,
                             values_only=True))
    return pd.DataFrame(data, columns=cols)

def _col_series(df: pd.DataFrame, header: str) -> pd.Series:
    idx = VM_COL_IDX[header] - 1
    if idx < df.shape[1]:
        return _to_number(df.iloc[:, idx])
    matches = [i for i,h in enumerate(df.columns) if h==header]
    if matches:
        log.warning("Header '%s' found at col %d, expected %d",
                    header, matches[0]+1, idx+1)
        return _to_number(df.iloc[:, matches[0]])
    raise RuntimeError(f"Header '{header}' not found in DataFrame")

def _make_block(df: pd.DataFrame, names: List[str]) -> pd.DataFrame:
    return pd.concat([_col_series(df, n) for n in names], axis=1)

# -------------------------------------------------------------------- #
#  Cell styling helpers
# -------------------------------------------------------------------- #
FORMULA_FILL = PatternFill("solid","D9D9D9","D9D9D9")

def _add_grand_total(ws, first_num_col: int):
    ws.insert_rows(2)
    ws.cell(2,1,"Grand Total").font = Font(bold=True)
    for col in range(first_num_col, ws.max_column + 1):
        L = get_column_letter(col)
        c = ws.cell(2, col, f"=SUM({L}3:{L}{ws.max_row})")
        c.font = Font(bold=True)
        c.number_format = "$#,##0.00"
        c.alignment = Alignment(horizontal="right")
        c.fill = FORMULA_FILL

# -------------------------------------------------------------------- #
#  Extract Comparison block (CK vs VM)
# -------------------------------------------------------------------- #
def _extract_comparison_adjustments(wb) -> pd.DataFrame:
    if "PnL Pivot" not in wb.sheetnames:
        log.warning("PnL Pivot sheet not found – skipping CK-VM adjustments")
        return pd.DataFrame(columns=MANUAL_COLUMNS)

    ws_cmp = wb["PnL Pivot"]

    # locate title
    title_row = next((r for r in range(1, ws_cmp.max_row+1)
                      if str(ws_cmp.cell(r,1).value).strip()=="Comparison"),None)
    if title_row is None:
        log.warning("Comparison block not found – skipping")
        return pd.DataFrame(columns=MANUAL_COLUMNS)
    hdr_row = title_row + 1

    # header map
    header_map: Dict[str,int] = {}
    col = 2
    while True:
        v = ws_cmp.cell(hdr_row,col).value
        if v in (None,""):
            break
        header_map[str(v).strip()] = col
        col += 1

    period_name = ws_cmp["B2"].value
    rows=[]
    r = hdr_row + 1
    while True:
        basin = ws_cmp.cell(r,1).value
        if basin in (None,"","Grand Total"):
            break
        rec = {c:0 for c in MANUAL_COLUMNS}
        rec["Project Number"]=""
        rec["ENG BASIN R1"]=basin
        rec["Period Name"]=period_name
        rec["COMMENT"]="CK vs VM Adj"
        for fld in SUMMARY_FIELDS:
            if fld not in header_map:
                rec[fld]=0
                continue
            xl_col = header_map[fld]
            xl_letter = get_column_letter(xl_col)
            rec[fld] = ("=-'PnL Pivot'!" if fld in COST_FIELDS
                        else "='PnL Pivot'!") + f"{xl_letter}{r}"
        rows.append(rec)
        r += 1

    return pd.DataFrame(rows, columns=MANUAL_COLUMNS)

# -------------------------------------------------------------------- #
#  Append rows to manual grid
# -------------------------------------------------------------------- #
def _append_manual_rows(ws, df: pd.DataFrame, num_align: Alignment) -> None:
    for _, rec in df.iterrows():
        ws.append([rec[c] for c in MANUAL_COLUMNS])
        r = ws.max_row
        for j,col_name in enumerate(MANUAL_COLUMNS,start=1):
            cell = ws.cell(r,j)
            if col_name in SUMMARY_FIELDS:
                cell.number_format="$#,##0.00"
                cell.alignment=num_align
        rev_formula = f"=SUM({get_column_letter(5)}{r}:{get_column_letter(10)}{r})"
        var_formula = f"=SUM({get_column_letter(12)}{r}:{get_column_letter(18)}{r})"
        for formula,col in ((rev_formula,4),(var_formula,11)):
            c = ws.cell(r,col,formula)
            c.number_format="$#,##0.00"
            c.alignment=num_align
            c.fill=FORMULA_FILL

# -------------------------------------------------------------------- #
#  Main generator
# -------------------------------------------------------------------- #
def generate_project_vm_adj(path: str) -> None:

    wb   = _safe_load_workbook(path)
    wsdb = wb[DB_SHEET]
    log.info("Workbook opened: %s", path)

    # ------------ Database DF ---------------------------------------- #
    rows = wsdb.iter_rows(min_row=HEADER_ROW, values_only=True)
    dfdb = pd.DataFrame(rows, columns=next(rows))
    dfdb.columns = [str(c).strip() if isinstance(c,str) else c
                    for c in dfdb.columns]
    log.info("[1] Database rows=%d cols=%d", *dfdb.shape)

    # ------------ FX factors ----------------------------------------- #
    fx_cad = _to_number(dfdb.iloc[:,COL_FX_CAD-1]).dropna().iloc[-1]
    fx_au  = _to_number(dfdb.iloc[:,COL_FX_AUD-1]).dropna().iloc[-1]
    eng    = dfdb.iloc[:,COL_ENG_BASIN-1].astype(str).str.strip()
    conv   = pd.Series(1.0,index=dfdb.index)
    conv.loc[eng.eq("CA")] = 1/fx_cad
    conv.loc[eng.eq("AU")] = 1/fx_au
    log.info("[2] FX CAD=%s AUD=%s", fx_cad, fx_au)

    # ------------ CK sums (prior month) ------------------------------ #
    pivot_idx = dfdb.columns.get_loc(PIVOT_MARKER)
    prev_lbl  = (datetime.date.today().replace(day=1) -
                 datetime.timedelta(days=1)).strftime("%b-%y")
    mask_prev = dfdb[PIVOT_MONTH_COL].eq(prev_lbl)
    ck_sums = {
        fld: 0.0 if col is None else
              _to_number(dfdb.iloc[mask_prev.values,
                                   [i for i,h in enumerate(dfdb.columns)
                                    if h==col and i>pivot_idx][0]]).sum()
        for fld,col in CK_COLUMNS.items()
    }

    # ------------ VM grand totals ------------------------------------ #
    vm_totals = {
        fld: 0.0 if not cols else
              (_make_block(dfdb,cols).multiply(conv,axis=0)
                                     .sum(axis=1).sum())
        for fld,cols in VM_SUM_MAP.items()
    }

    # ------------ split masks ---------------------------------------- #
    proj = dfdb.iloc[:,COL_PROJ_NUM-1].apply(
        lambda x:f"{int(x):06}" if pd.notna(x) and isinstance(x,(int,float))
                               else str(x).strip())
    main_pads = set(pd.to_numeric(
        dfdb.iloc[:, COL_MAIN_PAD-1], errors="coerce")
                      .dropna()
                      .astype(int)
                      .astype(str))
    prev_pads = set(pd.to_numeric(
        dfdb.iloc[:, COL_PREV_PAD-1], errors="coerce")
                      .dropna()
                      .astype(int)
                      .astype(str))
    masks={
        "Current": proj.isin(main_pads)&proj.str.match(r"^\d{6}$"),
        "Previous":(~proj.isin(main_pads))&proj.isin(prev_pads)&proj.str.match(r"^\d{6}$"),
        "Unalloc": ~proj.str.match(r"^\d{6}$"),
        "Unass": (~proj.isin(main_pads|prev_pads))&proj.str.match(r"^\d{6}$"),
    }
    def _bucket(mask):
        return {fld:0.0 if not cols else
                (_make_block(dfdb.loc[mask],cols)
                 .multiply(conv.loc[mask],axis=0)
                 .sum(axis=1).sum())
                for fld,cols in VM_SUM_MAP.items()}
    vm_curr,vm_prev,vm_unal,vm_unas = map(_bucket,masks.values())
    log.info("[3] Buckets computed")

    # ------------------------------------------------------------------ #
    #  P. VM - Adjustments sheet
    # ------------------------------------------------------------------ #
    if "P. VM - Adjustments" in wb.sheetnames:
        wb.remove(wb["P. VM - Adjustments"])
    ws_adj = wb.create_sheet("P. VM - Adjustments")

    bold=Font(bold=True)
    hdr_align=Alignment(horizontal="center",wrap_text=True)
    num_align=Alignment(horizontal="right")

    # header row
    for j,h in enumerate(["Line"]+SUMMARY_FIELDS,start=3):
        c=ws_adj.cell(1,j,h); c.font, c.alignment = bold, hdr_align

    # CK totals (row 2) & VM totals (row 4)
    ws_adj.cell(2,3,"CK PnL").font=bold
    ws_adj.cell(4,3,"Project VM Total").font=bold
    for i,f in enumerate(SUMMARY_FIELDS,start=4):
        ws_adj.cell(2,i,ck_sums[f]).number_format="$#,##0.00"
        ws_adj.cell(4,i,vm_totals[f]).number_format="$#,##0.00"

    # bucket rows 6-9
    for r,(lbl,sums) in enumerate(
        [("↳ Current Projects",vm_curr),
         ("↳ Previous Projects",vm_prev),
         ("↳ Unallocated",vm_unal),
         ("↳ Unassigned (0)",vm_unas)],start=6):
        ws_adj.cell(r,3,lbl).font=bold
        for i,f in enumerate(SUMMARY_FIELDS,start=4):
            ws_adj.cell(r,i,sums[f]).number_format="$#,##0.00"

    # Necessary Adjustment (row 11) & Adjustment Subtotal (row 12)
    ADD_HEADERS = {
        "VARIABLE COST","PROP COST","TRUCK COST","CHEM COST","FUEL COST",
        "MAT COST","OTHER PAD COST","ALLOC VM COST","MISC COST"
    }
    for col in range(4, 4 + len(SUMMARY_FIELDS)):
        header = SUMMARY_FIELDS[col - 4]
        L = get_column_letter(col)

        if header in ADD_HEADERS:              # cost buckets
            formula = f"=-{L}2-{L}4"            #  -L2 + L4
        else:                                  # revenue buckets
            formula = f"={L}2-{L}4"            #   L2 - L4

        ws_adj.cell(11, col, formula).number_format = "$#,##0.00"
        
        c12=ws_adj.cell(12,col)  # will set after manual-grid row-start known
        c12.number_format="$#,##0.00"
        for row in (11,12):
            ws_adj.cell(row,col).fill=FORMULA_FILL
    ws_adj.cell(11,3,"Necessary Adjustment (CK-VM)").font=bold
    ws_adj.cell(12,3,"Adjustment Subtotal").font=bold

    # ----------- NEW ROWS 14 & 15 ------------------------------------ #
    ws_adj.cell(14,3,"Adjusted P. VM Total").font=bold
    ws_adj.cell(15,3,"CK Delta Check").font=bold

    for col in range(4,4+len(SUMMARY_FIELDS)):
        L=get_column_letter(col)
        # row 14: VM Total + Adjustment Subtotal
        ws_adj.cell(14,col,f"={L}4+{L}12").number_format="$#,##0.00"
        ws_adj.cell(14,col).fill=FORMULA_FILL

        # row 15: CK Delta Check (rev: minus, cost: plus)
        header=SUMMARY_FIELDS[col-4]
        op = "+" if header in ADD_HEADERS else "-"
        ws_adj.cell(15,col,f"={L}2{op}{L}14").number_format="$#,##0.00"
        ws_adj.cell(15,col).fill=FORMULA_FILL

    # ------------------------------------------------------------------ #
    #  Manual grid
    # ------------------------------------------------------------------ #
    MAN_LABEL_ROW   = 17
    MAN_HEADER_ROW  = 18
    MAN_DATA_START  = 19   # used in Adjustment Subtotal formula

    # update Adjustment Subtotal row-12 formula range now that row-start known
    for col in range(4,4+len(SUMMARY_FIELDS)):
        L=get_column_letter(col)
        ws_adj.cell(12,col,
            f"=SUM({L}{MAN_DATA_START}:{L}200)").fill=FORMULA_FILL

    # Transload pre-fill
    df_trans=_read_block(wsdb,"BF","BM")
    df_trans["Adjusted NC"]=_to_number(df_trans["Adjusted NC"])
    df_t=df_trans[df_trans["Account Desc"]=="PROPPANT TRANSLOADING"].copy()
    df_t["VariableCostSign"]=df_t["Adjusted NC"]
    rows=[]
    for _,r in df_t.iterrows():
        rows.append({
            "Project Number":r["Project Number"],
            "ENG BASIN R1":r["ENG BASIN R1"],
            "Period Name":r["Period Name"],
            **{f:0 for f in SUMMARY_FIELDS},
            "PROP COST":r["VariableCostSign"],
            "COMMENT":"Transload",
        })
    for basin,grp in df_t.groupby("ENG BASIN R1"):
        rows.append({
            "Project Number":"",
            "ENG BASIN R1":basin,
            "Period Name":grp["Period Name"].iat[0],
            **{f:0 for f in SUMMARY_FIELDS},
            "PROP COST":-grp["Adjusted NC"].sum(),
            "COMMENT":"Transload Correction",
        })
    df_prefill=pd.DataFrame(rows,columns=MANUAL_COLUMNS)

    # CK-VM comparison adjustments
    df_cmp_adj=_extract_comparison_adjustments(wb)
    df_manual=pd.concat([df_prefill,df_cmp_adj],ignore_index=True)

    # grid label & header
    ws_adj.cell(MAN_LABEL_ROW,1,"Project VM Manual Adjustments").font=bold
    for j,h in enumerate(MANUAL_COLUMNS,start=1):
        c=ws_adj.cell(MAN_HEADER_ROW,j,h)
        c.font=bold; c.alignment=hdr_align

    # append manual rows
    _append_manual_rows(ws_adj,df_manual,num_align)

    # column widths
    for col in range(1,len(MANUAL_COLUMNS)+1):
        ws_adj.column_dimensions[get_column_letter(col)].width=25

    # ------------------------------------------------------------------ #
    #  Detail sheets (unchanged logic)
    # ------------------------------------------------------------------ #
    df_vm=_read_block(wsdb,"AO","BD")
    df_vm.columns=[c.strip() if isinstance(c,str) else c for c in df_vm.columns]
    df_vm["FX"]=conv.values
    for hdr in VM_COL_IDX:
        if hdr in df_vm.columns:
            df_vm[hdr]=_to_number(df_vm[hdr])*df_vm["FX"]
        else:
            log.error("Detail slice missing header '%s'",hdr)

    detail_cols=["Project Number","ENG BASIN R1","Period Name"]+list(VM_COL_IDX.keys())
    masks["All"]=df_vm.index
    order=["All","Current","Previous","Unalloc","Unass"]
    sheet=lambda k:"P. VM - All USD" if k=="All" else f"P. VM - {k}"
    for key in order:
        sh=sheet(key)
        if sh in wb.sheetnames:
            wb.remove(wb[sh])
        ws_det=wb.create_sheet(sh)
        sub=df_vm.loc[masks[key],detail_cols]

        for j,cname in enumerate(detail_cols,start=1):
            c=ws_det.cell(1,j,cname); c.font=bold; c.alignment=hdr_align

        start_row=3
        for r,rec in enumerate(sub.itertuples(index=False),start=start_row):
            for j,val in enumerate(rec,start=1):
                ws_det.cell(r,j,val)

        last_row=ws_det.max_row
        for row in ws_det.iter_rows(min_row=start_row,max_row=last_row,
                                    min_col=4,max_col=len(detail_cols)):
            for cell in row:
                cell.number_format="$#,##0.00"
                cell.alignment=Alignment(horizontal="right")
        _add_grand_total(ws_det,4)

    # reorder tabs
    desired=[sheet(k) for k in order]+["P. VM - Adjustments"]
    wb._sheets.sort(key=lambda s:desired.index(s.title)
                    if s.title in desired else len(desired))

    wb.save(path)
    log.info("✅ Workbook updated at %s",path)

# -------------------------------------------------------------------- #
#  CLI
# -------------------------------------------------------------------- #
if __name__=="__main__":
    if len(sys.argv)!=2:
        print("Usage: python main.py <MonthDataFile.xlsx>")
        sys.exit(1)
    generate_project_vm_adj(sys.argv[1])

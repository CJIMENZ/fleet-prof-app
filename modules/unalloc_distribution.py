#!/usr/bin/env python3
# modules/unalloc_distribution.py
"""
Distribute unallocated Sand, Handling, Chemical, and Daily costs
----------------------------------------------------------------
Steps
1.  Pull *unallocated* lines from "P. VM - Unalloc" **and** the
    non-6-digit rows (the true unalloc-txn lines) in "P. VM - Adjustments".
    ➜  Combined DF = NUMERATOR   (printed & written)
2.  Pull activity metrics from Database/Main-Combo + Chem Cost totals
    from "P. VM - Current" (only projects that exist in Main-Combo).
    ➜  Basin-level METRICS table = DENOMINATOR (printed & written)
3.  Build allocation ratios per basin, detect "orphans" where the
    denominator is 0, sprinkle those across active basins (≠ CA),
    and print three debug tables:
        3-a  orphan costs
        3-b  orphan ratios
        3-c  final basin ratios            (printed & written)
4.  Copy Main-Combo and append the four allocated cost columns
    (Unalloc_Sand / _Handle / _Chem / _Daily).               (printed & written)

Every DF shown in the console is also dropped into the worksheet
in the same order, separated by a blank row and a bold section title.
"""

from __future__ import annotations
import datetime as dt, re, sys, traceback
from datetime import datetime, date
from pathlib import Path
from typing import List, Dict
import logging

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

logger = logging.getLogger(__name__)

# ── Formatting ─────────────────────────────────────────────────────────────
LIGHT_GRAY   = PatternFill(fill_type="solid", fgColor="DDDDDD")
HEADER_FILL  = PatternFill(fill_type="solid", fgColor="C0C0C0")
CURRENCY_FMT = '$#,##0.00'
NUMBER_FMT   = '#,##0.00'

# ── Column header normalisation ────────────────────────────────────────────
RENAME_MAP = {
    'ENG BASIN R1':              'LBRT BASIN',
    'Chemical and Gel cost':     'Chem Cost',
    'Mat and Containment Costs': 'Mat Cost',
    'Other Pad Costs':           'Other Pad Cost',
    'Allocation VM':             'Alloc VM Cost',
    # sometimes Current sheet header shows "Chemical cost"
    'Chemical cost':             'Chem Cost',
}

# map the ALL-CAPS headers coming from Adjustments to the mixed-case
# equivalents used elsewhere
_COL_MAP = {
    "PROP COST": "Prop Cost",
    "TRUCK COST": "Truck Cost",
    "CHEM COST": "Chem Cost",
    "FUEL COST": "Fuel Cost",
    "MAT COST": "Mat Cost",
    "OTHER PAD COST": "Other Pad Cost",
    "ALLOC VM COST": "Alloc VM Cost",
}

# map duplicate-suffix headers to canonical names
_CANON_MAP = {
    "PAD NO": "Pad No",
    "PAD START": "Pad Start",
    "PAD END": "Pad End",
    "MONTH_YEAR_START": "MONTH_YEAR_START"
}

def _normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize column headers by removing duplicate suffixes and standardizing names."""
    canon_map = {}
    for col in df.columns:
        u = col.upper()
        for prefix, canon in _CANON_MAP.items():
            if u.startswith(prefix):
                canon_map[col] = canon
                break
    if canon_map:
        df = df.rename(columns=canon_map)
    return df

def _standardise_cost_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename cost columns to a single convention."""
    df = _normalize_headers(df)
    new_cols = {c: _COL_MAP.get(c.strip().upper(), c) for c in df.columns}
    return df.rename(columns=new_cols)

# ═══════════════════════════════════════════════════════════════════════════
def _find_sheet_name(wb, keywords: List[str]) -> str:
    """Return the first sheet whose name contains *all* keywords (case-insensitive)."""
    for name in wb.sheetnames:
        lname = name.lower()
        if all(k.lower() in lname for k in keywords):
            return name
    raise KeyError(f"No sheet name contains keywords: {keywords!r}")

# -------------------------------------------------------------------------- 
def _read_pvm_body(workbook_path: str, sheet_name: str) -> pd.DataFrame:
    """
    Read any P. VM sheet whose *real* data begin in row 4 (header row 1,
    grand-total row 2, blank row 3). Returns a cleaned DataFrame containing
    only cost columns (any column whose header contains "rev" or "revenue"
    is dropped).
    """
    df = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        header=0,            # row 1 is header
        engine="openpyxl"
    )
    df = df.iloc[2:]        # drop GT + blank
    df = df.dropna(how="all")          # strip empty rows at bottom
    
    # Keep only columns that are *not* revenue related
    col_mask = ~df.columns.astype(str).str.contains(r"rev|revenue", case=False, regex=True)
    df = df.loc[:, col_mask]
    
    # Normalize & clean up
    df.columns = [str(c).strip() for c in df.columns]
    df.rename(columns=RENAME_MAP, inplace=True)
    
    # Forward-fill merged cells in LBRT BASIN column if it exists
    if "LBRT BASIN" in df.columns:
        df["LBRT BASIN"].ffill(inplace=True)
        
    return df.reset_index(drop=True)

# -------------------------------------------------------------------------- 
def _read_pvm_adjustments(workbook_path: str, sheet_name: str) -> pd.DataFrame:
    """
    Return a DataFrame containing **only cost columns** from the
    "P. VM - Adjustments" worksheet.

    All columns whose header matches "rev" or "revenue"
    (case-insensitive) are removed.  Useful identifiers like
    "Project Number", "LBRT BASIN", "Period Name", and any *Cost*
    buckets are retained.
    """
    df = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        header=17,           # Excel row 18
        engine="openpyxl"
    )
    df = df.dropna(how="all")
    
    # Keep only non-revenue columns
    mask = ~df.columns.str.contains(r"rev|revenue", case=False, regex=True)
    df = df.loc[:, mask]

    # Drop Variable / Misc / Comment columns ---------------------------------
    _DROP = {"VARIABLE COST", "MISC COST", "COMMENT"}
    df = df.drop(columns=[c for c in df.columns if c.strip().upper() in _DROP],
                 errors="ignore")

    # Remove rows whose Project Number is a *numeric* 6-digit code -----------
    if "Project Number" in df.columns:
        def _is_six_digit(val):
            if pd.isna(val):
                return False
            try:
                n = int(float(val))
                return 100000 <= n <= 999999
            except (ValueError, TypeError):
                return False
        
        df = df[~df["Project Number"].apply(_is_six_digit)]
    
    # Tidy up
    df.columns = df.columns.str.strip()
    for col in ("LBRT BASIN", "Period Name"):
        if col in df.columns:
            df[col] = df[col].ffill()          # avoid chained-assignment warning

    df = df.reset_index(drop=True)
    
    logger.debug("P. VM – Adjustments (filtered):\n%s", df)
    
    return df

# -------------------------------------------------------------------------- 
def _read_main_combo(workbook_path: str) -> pd.DataFrame:
    df = pd.read_excel(
        workbook_path,
        sheet_name="Database",
        header=1,            # Excel row 2
        usecols="A:P",
        engine="openpyxl"
    )
    df = df.dropna(how="all", subset=["Pad No"])
    df.columns = [c.strip() for c in df.columns]
    df.rename(columns={"PAD START":"Pad Start", "PAD END":"Pad End"}, inplace=True)
    return df.reset_index(drop=True)

# --------------------------------------------------------------------------
def _read_stragglers(workbook_path: str,
                     main_combo_cols: pd.Index) -> pd.DataFrame:
    """
    Extract the straggler block by *position* (duplicate header set
    starting at column 18) instead of by fuzzy header names.
    """
    df_full = pd.read_excel(workbook_path, sheet_name="Database",
                            header=1, engine="openpyxl")

    # -- find second occurrence of 'Pad No' ------------------------------
    pad_cols = [i for i, c in enumerate(df_full.columns)
                if str(c).upper().startswith("PAD NO")]
    if len(pad_cols) < 2:
        logger.info("Straggler header set not found.")
        return pd.DataFrame()

    start_idx = pad_cols[1]                 # second occurrence
    col_span  = 11                          # see database_map.txt
    strag_df  = df_full.iloc[:, start_idx : start_idx + col_span].copy()

    strag_df.columns = [
        "Pad No", "MONTH_YEAR_START", "Is Previous Months",
        "Pad Start", "Pad End", "Customer", "Pad Name",
        "CREW SS", "LBRT BASIN", "HHP-HRS", "SAND SHORT TNS"
    ]

    # zero-out activity we ignore
    strag_df["HHP-HRS"]        = 0
    strag_df["SAND SHORT TNS"] = 0

    # drop empty rows
    strag_df = strag_df.dropna(how="all", subset=["Pad No"]).reset_index(drop=True)

    # ensure dates survive the later NA check
    for col in ("Pad Start", "Pad End"):
        strag_df[col] = _coerce_excel_datetime(strag_df[col])

    # numeric placeholders
    for c in ["Avg. Client Provided", "Prop TN", "Pump Time", "Pump Hours",
              "Non Sync Hours", "Sync Hours", "Chem Cost"]:
        strag_df[c] = 0

    strag_df = strag_df.reindex(columns=main_combo_cols)   # keeps the parsed dates
    return strag_df

# ═══════════════════════════════════════════════════════════════════════════
def _excel_serial_to_ts(s: pd.Series) -> pd.Series:
    """
    Convert Excel-serial floats in *s* to pandas Timestamps
    and leave everything else untouched.
    """
    num_mask = pd.to_numeric(s, errors="coerce").notna()   # element-wise
    converted = pd.to_datetime(s.where(num_mask),
                               unit="d", origin="1899-12-30", errors="coerce")
    return s.where(~num_mask, converted)                   # keep originals

def _coerce_excel_datetime(col: pd.Series) -> pd.Series:
    """
    Return a Series of pandas Timestamps.

    - ordinary date strings like '3/24/2025 04:30' are parsed
    - raw Excel serial floats/ints are converted (1899-12-30 origin)
    - anything un-convertible becomes NaT
    """
    # 1) first pass – try normal string parsing
    out = pd.to_datetime(col, errors="coerce")

    # 2) whatever is still NaT *might* be an Excel serial number
    mask_nat = out.isna() & col.notna()
    if mask_nat.any():
        serial = pd.to_numeric(col[mask_nat], errors="coerce")
        out.loc[mask_nat & serial.notna()] = pd.to_datetime(
            serial.dropna(), unit="d", origin="1899-12-30"
        )

    return out

def run_unalloc_distribution(
        workbook_path: str,
        month_start: date,
        month_end:   date
) -> None:

    print(f"\n[INFO] ► Unalloc Distribution run for: {workbook_path}")
    print(f"       Window: {month_start} -> {month_end}\n")

    try:
        wb = load_workbook(workbook_path, data_only=True)   # keep open for write-back
    except Exception:
        traceback.print_exc(); sys.exit(1)

    # ── Pull ALL raw data first ───────────────────────────────────────────
    sheet_unalloc  = _find_sheet_name(wb, ["p. vm", "unalloc"])
    sheet_current  = _find_sheet_name(wb, ["p. vm", "current"])
    sheet_adj      = _find_sheet_name(wb, ["p. vm", "adjust"])
    sheet_unass    = _find_sheet_name(wb, ["p. vm", "unass"])

    df_unalloc_raw = _read_pvm_body(workbook_path, sheet_unalloc)
    df_current_raw = _read_pvm_body(workbook_path, sheet_current)
    df_adjust_raw  = _read_pvm_adjustments(workbook_path, sheet_adj)
    df_unass_raw   = _read_pvm_body(workbook_path, sheet_unass)
    df_main_raw    = _read_main_combo(workbook_path)
    df_strag_raw   = _read_stragglers(workbook_path, df_main_raw.columns)

    # ---- print raw pulls -------------------------------------------------
    _dbg(df_unalloc_raw, "P. VM – Unalloc  (raw)")
    _dbg(df_adjust_raw,  "P. VM – Adjustments (raw)")
    _dbg(df_unass_raw,   "P. VM – Unass (raw)")
    _dbg(df_current_raw, "P. VM – Current  (raw)")
    _dbg(df_main_raw,    "Main_Combo (raw)")
    _dbg(df_strag_raw,   "Stragglers (raw)")

    # ╔════════ STEP 1 ═══════════════════════════════════════════════════╗
    # Combine unallocated lines  (numerator)
    def _step1_build_combined(df_unalloc, df_adjust, df_unass) -> pd.DataFrame:
        """
        Concatenate **only** P. VM – Unalloc and the true-unalloc rows from
        P. VM – Adjustments, align headers, drop Project Number,
        and SUM by basin.
        """
        # 1) Normalise column names -------------------------------------------------
        dfs = []
        for df in (df_unalloc, df_adjust, df_unass):
            if "ENG BASIN R1" in df.columns:
                df = df.rename(columns={"ENG BASIN R1": "LBRT BASIN"})
            dfs.append(_standardise_cost_columns(df))

        combined = pd.concat(dfs, ignore_index=True, sort=False)

        # 2) Drop unwanted identifier ----------------------------------------------
        combined = combined.drop(columns=[c for c in combined.columns
                                           if c.strip().upper() == "PROJECT NUMBER"],
                                 errors="ignore")

        # 3) Group & sum by basin ---------------------------------------------------
        numeric_cols = combined.select_dtypes(include="number").columns
        combined = (
            combined
            .groupby("LBRT BASIN", dropna=False)[numeric_cols]
            .sum(min_count=1)                 # keep NaN if all NaN for that basin
            .reset_index()
        )

        # 4) DEBUG – print full DataFrame ------------------------------------------
        _dbg(combined, "STEP 1 ► COMBINED NUMERATOR (summed by basin)")

        return combined

    df_num = _step1_build_combined(df_unalloc_raw, df_adjust_raw, df_unass_raw)

    # ╔════════ STEP 2 ═══════════════════════════════════════════════════╗
    # Combine Main-Combo + Stragglers, then build denominator metrics
    df_main = pd.concat(
        [df_main_raw, df_strag_raw],
        ignore_index=True,
        sort=False
    )

    _dbg(
        df_main,   # show full DataFrame instead of just tail
        "Main_Combo  +  Stragglers  (concatenated)"
    )

    # ----- continue with the date conversions & pad-day calc -------------
    for col in ("Pad Start", "Pad End"):
        df_main[col] = _coerce_excel_datetime(df_main[col])
    ms, me = pd.Timestamp(month_start), pd.Timestamp(month_end)

    mask = df_main["Pad Start"].isna() & df_main["Pad End"].isna()
    df_main.loc[mask, ["Pad Start", "Pad End"]] = [ms, me]

    # keep originals intact ⤵
    orig_start = df_main["Pad Start"].copy()
    orig_end   = df_main["Pad End"].copy()

    start_clip = orig_start.clip(lower=ms)
    end_clip   = orig_end  .clip(upper=me + pd.Timedelta(days=1) - pd.Timedelta(seconds=1))

    tdelta = (end_clip - start_clip).clip(lower=pd.Timedelta(0))
    df_main["pad_days"] = tdelta.dt.total_seconds().fillna(0) / 86_400

    # Convenience copy so you can eyeball it in Excel (last column)
    df_main["Pad_Days_Calc"] = df_main["pad_days"]

    # Chem cost (only pads present in Main-Combo)
    df_current = df_current_raw.copy()
    
    # --- make absolutely sure both keys are the same dtype -----------------
    df_current["Project Number"] = (
        pd.to_numeric(df_current["Project Number"], errors="coerce")
          .astype("Int64")           # nullable int so NaNs survive
    )

    df_main["Pad No"] = (
        pd.to_numeric(df_main["Pad No"], errors="coerce")
          .astype("Int64")
    )

    # clean chem column while we're here
    df_current["Chem Cost"] = (
        pd.to_numeric(df_current["Chem Cost"], errors="coerce")
          .fillna(0)
    )

    chem_by_pad = (
        df_current
        .groupby("Project Number")["Chem Cost"]
        .sum()
    )

    df_main["Chem Cost"] = (
        df_main["Pad No"]
          .map(chem_by_pad)
          .fillna(0)
    )

    # ── NEW: guard + weighted-proppant helper ─────────────
    if "Avg. Client Provided" not in df_main.columns:
        raise KeyError(
            "Column 'Avg. Client Provided' is missing in Main-Combo; "
            "can't compute PropTotal = Prop TN × Avg. Client Provided."
        )

    # multiply once, then group
    df_main["PropXClient"] = (
        df_main["Prop TN"].fillna(0) *
        df_main["Avg. Client Provided"].fillna(0)
    )

    # Basin-level denominator
    grp_m      = df_main.groupby("LBRT BASIN")
    prop_total = grp_m["PropXClient"].sum()          # ← changed line
    day_total  = grp_m["pad_days"].sum()
    chem_total = grp_m["Chem Cost"].sum()

    df_den = pd.DataFrame({
        "Basin":     prop_total.index,
        "PropTotal": prop_total.values,              # keep same column name
        "DayTotal":  day_total.reindex(prop_total.index, fill_value=0).values,
        "ChemTotal": chem_total.reindex(prop_total.index, fill_value=0).values
    })

    _dbg(df_den, "STEP 2 ► DENOMINATOR (Metrics)")

    # ╔════════ STEP 3 ═══════════════════════════════════════════════════╗
    # Build unalloc totals by basin  (numerator   ↓)
    grp_u  = df_num.groupby("LBRT BASIN")
    sand_u = grp_u["Prop Cost"].sum()
    hand_u = grp_u["Truck Cost"].sum()
    daily_u = (
        df_num[["Fuel Cost","Mat Cost","Other Pad Cost","Alloc VM Cost"]]
        .sum(axis=1)
        .groupby(df_num["LBRT BASIN"]).sum()
    )
    chem_u = grp_u["Chem Cost"].sum()

    # Prepare aligned series
    basins = df_den["Basin"].unique()
    def _al(s):
        return s.reindex(basins, fill_value=0)

    sand_u, hand_u, daily_u, chem_u = map(_al, (sand_u, hand_u, daily_u, chem_u))
    prop_total, day_total, chem_total = map(_al, (prop_total, day_total, chem_total))

    # -- orphan helper ----------------------------------------------------
    def _ratio(unalloc, denom):
        with pd.option_context("mode.use_inf_as_na", True):
            r = (unalloc / denom).fillna(0)
        return r

    def _sprinkle(unalloc, denom, base_ratio):
        """Return final ratio after sprinkling orphans (denom == 0)."""
        zero_mask = (denom == 0) & (unalloc > 0)
        orphan_cost = unalloc[zero_mask].sum()
        valid_mask  = (denom > 0) & (base_ratio.index != "CA")
        pool = denom[valid_mask].sum()
        orphan_ratio = orphan_cost / pool if pool else 0
        final = base_ratio.copy()
        final.loc[valid_mask] += orphan_ratio
        return orphan_cost, orphan_ratio, final

    # -- Compute base ratios ---------------------------------------------
    ratio_sand   = _ratio(sand_u,  prop_total)
    ratio_handle = _ratio(hand_u,  prop_total)
    ratio_daily  = _ratio(daily_u, day_total)
    ratio_chem   = _ratio(chem_u,  chem_total)

    # -- Sprinkle ---------------------------------------------------------
    orphan_sand,   orph_r_sand,   final_sand   = _sprinkle(sand_u,  prop_total, ratio_sand)
    orphan_handle, orph_r_handle, final_handle = _sprinkle(hand_u,  prop_total, ratio_handle)
    orphan_daily,  orph_r_daily,  final_daily  = _sprinkle(daily_u, day_total,  ratio_daily)
    orphan_chem,   orph_r_chem,   final_chem   = _sprinkle(chem_u,  chem_total, ratio_chem)

    # -- Debug tables -----------------------------------------------------
    df_orphans = pd.DataFrame({
        "Metric": ["Sand","Handle","Daily","Chem"],
        "OrphanCost": [orphan_sand, orphan_handle, orphan_daily, orphan_chem]
    })
    _dbg(df_orphans, "STEP 3-a ► ORPHAN COSTS")

    df_orphan_ratio = pd.DataFrame({
        "Metric": ["Sand","Handle","Daily","Chem"],
        "OrphanRatio": [orph_r_sand, orph_r_handle, orph_r_daily, orph_r_chem]
    })
    _dbg(df_orphan_ratio, "STEP 3-b ► ORPHAN RATIOS  (added to active basins ≠ CA)")

    df_ratios = pd.DataFrame({
        "Basin": final_sand.index,
        "RatioSand":   final_sand.values,
        "RatioHandle": final_handle.values,
        "RatioDaily":  final_daily.values,
        "RatioChem":   final_chem.values
    })
    _dbg(df_ratios, "STEP 3-c ► FINAL BASIN RATIOS")

    # ╔════════ STEP 4 ═══════════════════════════════════════════════════╗
    # Allocate per-pad
    df_out = df_main.copy()
    df_out["Unalloc_Sand"]   = df_out["Prop TN"]   * df_out["LBRT BASIN"].map(final_sand)
    df_out["Unalloc_Handle"] = df_out["Prop TN"]   * df_out["LBRT BASIN"].map(final_handle)
    df_out["Unalloc_Chem"]   = df_out["Chem Cost"] * df_out["LBRT BASIN"].map(final_chem)
    df_out["Unalloc_Daily"]  = df_out["pad_days"]  * df_out["LBRT BASIN"].map(final_daily)

    _dbg(df_out.head(), "STEP 4 ► PAD-LEVEL ALLOCATIONS (first rows)")

    # ╔════════ WRITE TO EXCEL ═══════════════════════════════════════════╗
    if "Unalloc_Distribution" in wb.sheetnames:
        del wb["Unalloc_Distribution"]
    ws = wb.create_sheet("Unalloc_Distribution")

    def _write_section(title: str, df: pd.DataFrame, start_row: int) -> int:
        """Write a DF preceded by a bold title, return next free row index."""
        cell = ws.cell(start_row, 1, title)
        cell.font = Font(bold=True, size=12)
        cell.fill = HEADER_FILL
        start_row += 2
        for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                # ➋ normalise *every* timestamp that comes through here
                if isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()          # <-- plain datetime
                if isinstance(value, datetime):
                    ws.cell(r_idx, c_idx, value).number_format = "yyyy-mm-dd hh:mm"
                else:
                    ws.cell(r_idx, c_idx, value)
                # header row styling
                if r_idx == start_row:
                    cell = ws.cell(r_idx, c_idx, value)
                    cell.font = Font(bold=True)
                    cell.fill = HEADER_FILL
                # Currency formatting heuristics
                if isinstance(value, (int,float)) and ("Cost" in df.columns[c_idx-1] or "Unalloc" in df.columns[c_idx-1]):
                    cell = ws.cell(r_idx, c_idx, value)
                    cell.number_format = CURRENCY_FMT
        return r_idx + 2   # blank row after table

    row = 1
    row = _write_section("RAW – P. VM Unalloc",      df_unalloc_raw, row)
    row = _write_section("RAW – P. VM Adjustments",  df_adjust_raw,  row)
    row = _write_section("RAW – P. VM Unass",        df_unass_raw,   row)
    row = _write_section("RAW – P. VM Current",      df_current_raw, row)
    row = _write_section("RAW – Main_Combo",         df_main,    row)
    row = _write_section("RAW – Stragglers",         df_strag_raw, row)
    row = _write_section("STEP 1 – Numerator Combined", df_num,      row)
    row = _write_section("STEP 2 – Denominator Metrics", df_den,     row)
    row = _write_section("STEP 3-a – Orphan Costs",      df_orphans, row)
    row = _write_section("STEP 3-b – Orphan Ratios",     df_orphan_ratio, row)
    row = _write_section("STEP 3-c – Final Basin Ratios", df_ratios,  row)
    _ = _write_section("STEP 4 – Pad-level Allocations", df_out, row)

    wb.save(workbook_path)
    print(f"[INFO] ✔ Unalloc_Distribution sheet written & workbook saved\n")

# ─── Pretty-print full frames ────────────────────────────────────────
def _dbg(df: pd.DataFrame, tag: str) -> None:
    """Console dump of *entire* DataFrame (no head truncation)."""
    with pd.option_context(
        "display.max_rows", None,
        "display.max_columns", None,
        "display.width",     None,        # pandas won't fold columns
        "display.float_format", "{:,.6g}".format,
    ):
        print(f"[DEBUG] {tag}: shape={df.shape}")
        print(df.to_string(index=False), "\n")

# ── CLI helper ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python -m modules.unalloc_distribution "
              "<workbook_path> <YYYY-MM-DD start> <YYYY-MM-DD end>")
        sys.exit(1)
    run_unalloc_distribution(
        sys.argv[1],
        date.fromisoformat(sys.argv[2]),
        date.fromisoformat(sys.argv[3])
    )
